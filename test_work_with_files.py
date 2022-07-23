import os
import zipfile
import csv
import pytest

from openpyxl import load_workbook
from PyPDF2 import PdfReader


@pytest.fixture(scope="module", autouse=True)
def create_zip_archive():
    overall_zip = zipfile.ZipFile('resources/overall_zip.zip', 'w')

    for folder, subfolders, files in os.walk('resources/raw_resources'):
        for file in files:
            if file.endswith('.xlsx') or file.endswith('.pdf') or file.endswith('.csv'):
                overall_zip.write(
                    os.path.join(folder, file),
                    os.path.relpath(os.path.join(file), 'resources/test_resources'),
                    compress_type=zipfile.ZIP_DEFLATED
                )
    overall_zip.close()


def test_zip_archive_creation():
    zip_file = os.path.abspath('resources/overall_zip.zip')
    assert os.path.exists(zip_file)


class TestZIPArchive:
    @pytest.fixture(scope="class", autouse=True)
    def unzip_archive(self):
        unpack_zip = zipfile.ZipFile('resources/overall_zip.zip')
        unpack_zip.extractall('resources/test_resources')
        unpack_zip.close()

    def test_read_pdf_file(self):
        reader = PdfReader('resources/test_resources/sample.pdf')
        page = reader.pages[0]
        text = page.extract_text()
        assert 'demonstration' in text

    def test_read_xlsx_file(self):
        workbook = load_workbook('resources/test_resources/file_example_50.xlsx')
        sheet = workbook.active
        name = sheet.cell(column=2, row=20,).value
        assert 'Loreta' == name

    def test_read_csv_file(self):
        with open('resources/test_resources/username.csv') as f:
            reader = csv.reader(f)
            headers = next(reader)
        assert 'Identifier' in str(headers)


@pytest.fixture(scope="session", autouse=True)
def delete_archive():
    yield
    for file in [f for f in os.listdir('resources/test_resources')]:
        os.remove(os.path.join('resources/test_resources', file))

    zip_file = os.path.abspath('resources/overall_zip.zip')
    os.remove(zip_file)
